import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "求職トラッカー"

# ---- カラー定義 ----
RED    = "C0392B"
ORANGE = "E67E22"
GRAY   = "7F8C8D"
WHITE  = "FFFFFF"
LIGHT_RED    = "FADBD8"
LIGHT_ORANGE = "FDEBD0"
LIGHT_GRAY   = "F2F3F4"
HEADER_BG    = "2C3E50"

# ---- データ ----
sections = [
    {
        "label": "優先度：高",
        "color": RED,
        "bg": LIGHT_RED,
        "rows": [
            ("ソニーグループ", "事業企画・国際BD・PM", "新卒/転職", "検討中", "", "2027新卒ポータル確認・PS/SIE部門の求人調査"),
            ("安川電機", "国際営業・中国向けBD", "中途推奨", "検討中", "", "九州大学キャリアイベントで接点づくり・OB訪問"),
            ("SoftBankロボティクス", "BD・PM（ABB統合後）", "転職", "未検討", "", "ABB統合完了（2026年中）後に採用ページを再確認"),
            ("ファナック", "海外営業・中国語担当BD", "中途推奨", "検討中", "", "中途採用ページの求人内容確認・製造業知識を補強"),
        ]
    },
    {
        "label": "優先度：中",
        "color": ORANGE,
        "bg": LIGHT_ORANGE,
        "rows": [
            ("Preferred Networks", "PM・マーケティング・BD", "転職/新卒", "検討中", "", "kachaka関連の非エンジニア職が出たら即確認"),
            ("日本NVIDIA", "Partner BD・Field Marketing", "転職", "未検討", "", "中期ターゲット。先に業界実績を積んでから再検討"),
            ("クアルコムジャパン", "IoT BD・プロダクトマーケティング", "転職", "未検討", "", "中期ターゲット。自動車IoT求人をウォッチ"),
        ]
    },
    {
        "label": "優先度：低（長期監視）",
        "color": GRAY,
        "bg": LIGHT_GRAY,
        "rows": [
            ("ラピダス", "海外顧客向けBD（将来）", "転職", "未検討", "", "2027年量産フェーズ移行後に再評価"),
            ("ソシオネクスト", "FAE・海外営業", "新卒/転職", "未検討", "", "非エンジニア求人が出た場合のみ検討"),
            ("アームジャパン", "Partner BD（将来）", "転職", "未検討", "", "半導体業界実績を積んだ後の長期ターゲット"),
            ("ASMLジャパン", "カスタマーリレーション・オペレーション", "転職", "未検討", "", "非エンジニア職の求人が出たら確認"),
        ]
    },
]

headers = ["企業名", "ターゲット職種", "採用形態", "ステータス", "応募日", "次のアクション"]

# ---- ヘッダー行 ----
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = Font(bold=True, color=WHITE, size=11)
    cell.fill = PatternFill("solid", fgColor=HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

ws.row_dimensions[1].height = 24

# ---- データ行 ----
current_row = 2
for section in sections:
    # セクション見出し行
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    label_cell = ws.cell(row=current_row, column=1, value=section["label"])
    label_cell.font = Font(bold=True, color=WHITE, size=10)
    label_cell.fill = PatternFill("solid", fgColor=section["color"])
    label_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    label_cell.border = border
    ws.row_dimensions[current_row].height = 20
    current_row += 1

    for row_data in section["rows"]:
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor=section["bg"])
            cell.alignment = Alignment(vertical="center", wrap_text=True,
                                       horizontal="center" if col in (3, 4, 5) else "left")
            cell.font = Font(size=10)
            cell.border = border
        ws.row_dimensions[current_row].height = 32
        current_row += 1

# ---- 列幅 ----
col_widths = [22, 30, 14, 12, 12, 44]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ---- ウィンドウ枠固定（ヘッダー固定） ----
ws.freeze_panes = "A2"

# ---- 保存 ----
out = "/Users/yuezixuan/Desktop/CareerCollection/求職トラッカー.xlsx"
wb.save(out)
print(f"saved: {out}")
